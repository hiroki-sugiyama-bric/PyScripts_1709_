import os
import openpyxl as px
from openpyxl.styles import PatternFill

work_root_dir = '/Users/hirokisugiyama/Work/NTTTX/NTTTX_201709_/tests/utils/py3/all'
# py3_wb_path = os.path.join(work_root_dir, 'form_3_all_cp932.xlsx')
py3_wb_path = os.path.join(work_root_dir, 'form_3_html5lib_cp932.xlsx')
# label_wbs_dir = os.path.join(work_root_dir, '2_labeled_excels')
label_wbs_dir = os.path.join(work_root_dir, 'label_html_parser')

py3_ws_name = 'form_with_label'
yellow_fill = PatternFill(
    start_color='FFFFFACD',
    end_color='FFFFFACD',
    fill_type='solid'
)
label_columns = 'HIJKLM'

def get_labels_wb_paths():
    files = os.listdir(label_wbs_dir)
    excels = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$')]
    wb_paths = [os.path.join(label_wbs_dir, e) for e in excels]

    return wb_paths

def extract_id_attrs(row):
    website = transaction = form = None
    for cell in row:
        if cell.column == 'A':
            website = cell.value
            # 番号は違う場合がある
            if '.' in website:
                website = website[website.find('.') + 1:]

        elif cell.column == 'B':
            transaction = cell.value
        elif cell.column == 'C':
            form = cell.value

    return website, transaction, form

def create_rows_map(ws):
    rows_map = {extract_id_attrs(r): r for r in ws.rows}

    return rows_map

def find_cell(row, column_name):
    for cell in row:
        if cell.column == column_name:
            return cell

    return None

def copy_row_vals(src_row, dest_row, columns):
    for col in columns:
        src_cell = find_cell(src_row, col)
        dest_cell = find_cell(dest_row, col)

        dest_cell.value = src_cell.value

def mark_labeled_row(row):
    for cell in row:
        if cell.column in label_columns:
            cell.fill = yellow_fill

def migrate_labels():
    py3_wb = px.load_workbook(py3_wb_path)
    py3_ws = py3_wb.get_sheet_by_name(py3_ws_name)
    py3_rows_map = create_rows_map(py3_ws)

    labeled_paths = get_labels_wb_paths()
    for path in labeled_paths:
        labeled_wb = px.load_workbook(path)
        ws_name = labeled_wb.get_sheet_names()[0]
        labeled_ws = labeled_wb.get_sheet_by_name(ws_name)
        labeled_rows_map = create_rows_map(labeled_ws)

        for id_attrs, py3_row in py3_rows_map.items():
            labeled_row = labeled_rows_map.get(id_attrs, None)
            # 対応する行があれば処理する
            if labeled_row is not None:
                # ラベルデータが存在したことを示すために色を付ける
                mark_labeled_row(py3_row)
                # ラベルデータをコピー
                copy_row_vals(labeled_row, py3_row, label_columns)

    py3_wb.save(py3_wb_path[:-len('.xlsx')] + '_migrated.xlsx')


if __name__ == '__main__':
    migrate_labels()



